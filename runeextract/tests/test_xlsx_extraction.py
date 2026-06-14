"""Tests for XLSX full extraction."""

import os
import tempfile
from runeextract import extract


def test_xlsx_extraction():
    """Test basic XLSX text extraction."""
    from openpyxl import Workbook
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        path = tmp.name
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Name"
        ws["B1"] = "Age"
        ws["A2"] = "Alice"
        ws["B2"] = 30
        wb.save(path)

        result = extract(path)
        assert result.source_type == "xlsx"
        assert "Alice" in result.text
        assert len(result.tables) >= 1
        assert result.metadata.get("sheet_count", 0) >= 1
    finally:
        if os.path.exists(path):
            os.unlink(path)


def test_xlsx_multiple_sheets():
    """Test XLSX with multiple sheets."""
    from openpyxl import Workbook
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        path = tmp.name
    try:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "SheetA"
        ws1["A1"] = "X"
        ws2 = wb.create_sheet("SheetB")
        ws2["A1"] = "Y"

        wb.save(path)

        result = extract(path)
        assert result.metadata.get("sheet_count", 0) >= 2
        assert "SheetA" in result.text or "SheetB" in result.text or "X" in result.text or "Y" in result.text
        sheets = result.metadata.get("sheet_names", [])
        assert "SheetA" in sheets
        assert "SheetB" in sheets
    finally:
        if os.path.exists(path):
            os.unlink(path)


def test_xlsx_with_metadata():
    """Test XLSX metadata extraction."""
    from openpyxl import Workbook
    from openpyxl.packaging.core import DocumentProperties
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        path = tmp.name
    try:
        wb = Workbook()
        wb.active["A1"] = "data"
        wb.properties.title = "Test XLSX"
        wb.properties.creator = "Test Author"
        wb.save(path)

        result = extract(path, metadata=True)
        assert result.metadata.get("title") == "Test XLSX"
    finally:
        if os.path.exists(path):
            os.unlink(path)
