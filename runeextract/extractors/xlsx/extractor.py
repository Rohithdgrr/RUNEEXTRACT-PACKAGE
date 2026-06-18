"""
XLSX extractor using openpyxl.
"""

import logging
from openpyxl import load_workbook
from typing import List, Dict, Any
from io import BytesIO
from runeextract.core.extractor import BaseExtractor
from runeextract.models.document import Document as RuneDocument, Table
from runeextract.exceptions import ExtractionError, WrongPasswordError

logger = logging.getLogger(__name__)


class XlsxExtractor(BaseExtractor):
    """Extractor for XLSX files."""

    def extract(self, file_path: str) -> RuneDocument:
        self.validate_file(file_path)
        password = self.options.get('password')
        if password:
            wb = self._open_protected(file_path, password)
        else:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            text_parts: List[str] = []
            tables: List[Table] = []
            metadata = self._extract_metadata(wb)
            metadata['sheet_count'] = len(wb.sheetnames)
            metadata['sheet_names'] = wb.sheetnames

            for sheet_index, sheet_name in enumerate(wb.sheetnames, start=1):
                ws = wb[sheet_name]
                text_parts.append(f"\n--- Sheet: {sheet_name} ---\n\n")
                rows_data = []
                if ws.max_row and ws.max_column:
                    for row in ws.iter_rows(values_only=True):
                        row_values = [str(cell) if cell is not None else "" for cell in row]
                        rows_data.append(row_values)
                if rows_data:
                    columns = rows_data[0] if rows_data else []
                    data_rows = rows_data[1:]
                    has_data = any(any(cell for cell in row) for row in data_rows)
                    if has_data:
                        tables.append(Table(rows=data_rows, columns=columns,
                                            metadata={'sheet_name': sheet_name, 'sheet_index': sheet_index,
                                                      'max_row': len(data_rows), 'max_column': len(columns)}))
                    for row in rows_data:
                        text_parts.append('\t'.join(row))
                        text_parts.append('\n')
                    text_parts.append('\n')

            text = self.clean_text("".join(text_parts))
        finally:
            wb.close()
        return RuneDocument(text=text, tables=tables, images=[], metadata=metadata,
                            source_type="xlsx", source_path=file_path)

    @staticmethod
    def _open_protected(file_path: str, password: str):
        try:
            import msoffcrypto
        except ImportError:
            raise ExtractionError(
                "Password-protected XLSX requires msoffcrypto-tool. Install: pip install msoffcrypto-tool",
                file_path=file_path, error_code="E004"
            )
        with open(file_path, "rb") as f:
            decrypted = BytesIO()
            office_file = msoffcrypto.OfficeFile(f)
            try:
                office_file.load_key(password=password)
            except Exception:
                raise WrongPasswordError(file_path)
            office_file.decrypt(decrypted)
        decrypted.seek(0)
        return load_workbook(decrypted, read_only=True, data_only=True)

    def _extract_metadata(self, wb) -> Dict[str, Any]:
        metadata = {}
        try:
            props = wb.properties
            if props:
                metadata['title'] = props.title or ""
                metadata['creator'] = props.creator or ""
                metadata['description'] = props.description or ""
                metadata['keywords'] = props.keywords or ""
                metadata['created'] = str(props.created) if props.created else ""
                metadata['modified'] = str(props.modified) if props.modified else ""
                metadata['category'] = props.category or ""
        except Exception as exc:
            logger.debug(f"Metadata extraction error in XLSX: {exc}")
        return metadata

    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xls"]
